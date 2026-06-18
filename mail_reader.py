import imaplib
import email
import re
import os
import time
from openpyxl import load_workbook

# --- CONFIGURATION ---
IMAP_SERVER = "imap.gmail.com"
EMAIL_USER = "nouijai.ghassane@gmail.com"
EMAIL_PASS = "smstybvygospvxrs"
EXCEL_PATH = r"C:\Users\nouij\OneDrive\Internship\Automatica\Dashboard_Global Affairs Team_VBA.xlsm"

def parse_email_body(text_content):
    """Extracts raw key-value submission rows out of the payload body."""
    data = {}
    markers = ['TYPE', 'ACTION', 'CATEGORY', 'PRIORITY', 'DUE_DATE', 'OWNER', 'COMMENTS']
    for marker in markers:
        match = re.search(f"{marker}:(.*)", text_content)
        data[marker] = match.group(1).strip() if match else ""
    return data

def append_to_excel(task_data):
    """Locates target worksheets and safely appends incoming tasks bottom-up."""
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: Target file path not found: {EXCEL_PATH}")
        return False
        
    try:
        wb = load_workbook(EXCEL_PATH, keep_vba=True)
        
        # Decide target sheet based on action type
        if task_data['TYPE'] == "NEW_TASK":
            ws = wb["BackEnd"]
        elif task_data['TYPE'] == "COMPLETE_TASK":
            ws = wb["Finished"]
        else:
            return False

        # Find the next open empty row at the bottom of Column A
        next_row = 2
        while ws.cell(row=next_row, column=1).value is not None:
            next_row += 1
            
        # Calculate next incremental Index ID number
        prev_id = ws.cell(row=next_row-1, column=1).value if next_row > 2 else 0
        try:
            next_id = int(prev_id) + 1
        except:
            next_id = 1

        # Map values exactly to your original column schema
        ws.cell(row=next_row, column=1, value=next_id)                   # ID (Col A)
        ws.cell(row=next_row, column=4, value=task_data['ACTION'])       # Action (Col D)
        ws.cell(row=next_row, column=5, value=task_data['CATEGORY'])     # Category (Col E)
        ws.cell(row=next_row, column=6, value=task_data['PRIORITY'])     # Priority (Col F)
        ws.cell(row=next_row, column=7, value=task_data['DUE_DATE'])     # Due Date (Col G)
        ws.cell(row=next_row, column=9, value=task_data['OWNER'])        # Owner (Col I)
        ws.cell(row=next_row, column=11, value=task_data['COMMENTS'])    # Comments (Col K)

        wb.save(EXCEL_PATH)
        wb.close()
        print(f"Successfully synced {task_data['TYPE']} to row {next_row}!")
        return True
    except Exception as e:
        print(f"Failed to write to Excel: {e}")
        return False

def check_and_sync_mail():
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select("inbox")

        # Query for unread synchronization signals
        status, response = mail.search(None, '(UNSEEN SUBJECT "DATA_SYNC:")')
        if status != "OK":
            return

        email_ids = response[0].split()
        if not email_ids:
            return

        print(f"Found {len(email_ids)} pending tasks to synchronize...")

        for e_id in email_ids:
            status, data = mail.fetch(e_id, '(RFC822)')
            if status != "OK":
                continue
                
            raw_email = data[0][1]
            msg = email.message_from_bytes(raw_email)
            
            # Extract plain text content body
            body = ""
            if msg.is_multipart():
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                        break
            else:
                body = msg.get_payload(decode=True).decode('utf-8', errors='ignore')

            if "START_DATA" in body:
                parsed_task = parse_email_body(body)
                if append_to_excel(parsed_task):
                    # Mark as read only after successful Excel write
                    mail.store(e_id, '+FLAGS', '\\Seen')
                    
        mail.logout()
    except Exception as e:
        print(f"Network Connection Loop Error: {e}")

if __name__ == "__main__":
    print("Local OneDrive Synchronization Broker is running...")
    print("Press CTRL+C to terminate application.")
    while True:
        check_and_sync_mail()
        time.sleep(30)  # Checks your inbox cleanly every 30 seconds